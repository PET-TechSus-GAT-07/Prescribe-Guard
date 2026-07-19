#!/usr/bin/env python3
"""Importa o dataset RENAME (planilha local, UpToDate) para o SQLite do Prescribe-Guard.

Fonte: planilha `RENAME.xlsx`, mantida fora do repositório (não versionada),
com duas abas: a lista oficial de medicamentos essenciais ("RENAME (original)")
e as interações medicamentosas levantadas manualmente a partir do UpToDate
("RENAME (com IM)"). O texto clínico (mecanismo e manejo do paciente) é
gravado como veio da planilha, sem reescrita.

Import ADITIVO: não apaga nada do banco. Deve rodar DEPOIS de build_data.py,
porque build_data.py faz um replace total das tabelas medications/interactions
e apagaria o que este script adicionou.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
import unicodedata
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

from data_pipeline_common import (
    EFFECT_PATTERNS,
    SEVERITY_PRIORITY,
    SYSTEM_PATTERNS,
    canonical_pair,
    infer_labels,
    infer_severity,
    init_db,
    normalize_text,
    publish_frontend_database,
)


DEFAULT_DB_PATH = Path("storage/prescribe_guard.sqlite")
DEFAULT_PUBLIC_DB_PATH = Path("data/prescribe_guard.sqlite")
DEFAULT_SYNONYMS_PATH = Path("scripts/rename_synonyms.json")

SHEET_INTERACTIONS = "RENAME (com IM)"
SHEET_FORMULARY = "RENAME (original)"

RISK_TO_SEVERITY = {
    "a": "safe",
    "b": "minor",
    "c": "moderate",
    "d": "major",
    "x": "contraindicated",
}
RISK_PATTERN = re.compile(r"^([abcdx])\s*:", re.IGNORECASE)

ATC_GROUP_LABELS = {
    "A": "Trato alimentar e metabolismo",
    "B": "Sangue e órgãos hematopoiéticos",
    "C": "Sistema cardiovascular",
    "D": "Dermatológico",
    "G": "Sistema genito-urinário / hormônios sexuais",
    "H": "Hormônios sistêmicos (exceto sexuais)",
    "J": "Anti-infeccioso sistêmico",
    "L": "Antineoplásico / imunomodulador",
    "M": "Sistema musculoesquelético",
    "N": "Sistema nervoso",
    "P": "Antiparasitário",
    "R": "Sistema respiratório",
    "S": "Órgãos sensoriais",
    "V": "Diversos",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa o dataset RENAME (com IM) para o SQLite do Prescribe-Guard, de forma aditiva.",
        epilog=(
            "IMPORTANTE: rode este script DEPOIS de build_data.py. build_data.py faz um "
            "replace total do banco a partir da API remota e apagaria os dados do RENAME."
        ),
    )
    parser.add_argument("--xlsx", required=True, help="Caminho local do RENAME.xlsx.")
    parser.add_argument("--db", default=str(DEFAULT_DB_PATH), help="Caminho do banco SQLite local.")
    parser.add_argument(
        "--public-db",
        default=str(DEFAULT_PUBLIC_DB_PATH),
        help="Caminho do SQLite estático publicado para o frontend.",
    )
    parser.add_argument(
        "--synonyms",
        default=str(DEFAULT_SYNONYMS_PATH),
        help="Tabela de sinônimos de nomes (RENAME -> catálogo atual).",
    )
    return parser.parse_args()


def strip_accents_casefold(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    without_accents = "".join(c for c in normalized if unicodedata.category(c) != "Mn")
    return without_accents.casefold().strip()


def load_synonyms(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        strip_accents_casefold(key): value
        for key, value in raw.items()
        if not key.startswith("_")
    }


def title_case_dcb(name: str) -> str:
    name = normalize_text(name)
    if not name:
        return name
    return name[0].upper() + name[1:]


def resolve_medication_name(
    raw_name: str,
    synonyms: dict[str, str],
    existing_by_key: dict[str, str],
) -> str:
    key = strip_accents_casefold(raw_name)
    if key in existing_by_key:
        return existing_by_key[key]
    if key in synonyms:
        return synonyms[key]
    return title_case_dcb(raw_name)


def atc_category(atc_code: str) -> str:
    if not atc_code:
        return ""
    return ATC_GROUP_LABELS.get(atc_code[0].upper(), "")


def parse_risco(value: str) -> str | None:
    match = RISK_PATTERN.match(value.strip())
    if not match:
        return None
    return RISK_TO_SEVERITY.get(match.group(1).lower())


def read_formulary(workbook: Any) -> dict[str, dict[str, str]]:
    """DCB (chave normalizada) -> {display_name, atc_code, aware}.

    Linhas com DCB vazio herdam o DCB da linha anterior (célula mesclada no
    Excel entre diferentes formas farmacêuticas/concentrações do mesmo fármaco).
    """
    sheet = workbook[SHEET_FORMULARY]
    formulary: dict[str, dict[str, str]] = {}
    last_dcb_display = ""

    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        dcb_cell = normalize_text(row[0]) if row[0] else ""
        atc_cell = normalize_text(row[5]) if len(row) > 5 and row[5] else ""
        aware_cell = normalize_text(row[6]) if len(row) > 6 and row[6] else ""

        if dcb_cell:
            last_dcb_display = dcb_cell
        if not last_dcb_display:
            continue

        key = strip_accents_casefold(last_dcb_display)
        entry = formulary.setdefault(
            key, {"display_name": last_dcb_display, "atc_code": "", "aware": ""}
        )
        if atc_cell and not entry["atc_code"]:
            entry["atc_code"] = atc_cell
        if aware_cell and not entry["aware"]:
            entry["aware"] = aware_cell

    return formulary


def _interaction_score(record: dict[str, Any]) -> tuple[int, int, int, int]:
    return (
        len(record["mechanism"]) + len(record["recommendation"]),
        SEVERITY_PRIORITY[record["severity"]],
        len(record["effects"]),
        len(record["systems_affected"]),
    )


def read_interactions(
    workbook: Any,
    synonyms: dict[str, str],
    existing_by_key: dict[str, str],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    sheet = workbook[SHEET_INTERACTIONS]
    interactions_by_pair: dict[str, dict[str, Any]] = {}
    stats = {"rows_seen": 0, "rows_skipped_malformed": 0, "rows_skipped_empty": 0}

    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row or not row[0] or len(row) < 2 or not row[1]:
            continue
        stats["rows_seen"] += 1

        raw_dcb = normalize_text(row[0])
        raw_farmaco = normalize_text(row[1])
        risco = normalize_text(row[2]) if len(row) > 2 and row[2] else ""
        resumo = normalize_text(row[3]) if len(row) > 3 and row[3] else ""
        manejo = normalize_text(row[6]) if len(row) > 6 and row[6] else ""

        severity = parse_risco(risco)
        if risco and severity is None:
            # Linha com colunas deslocadas na origem (ex.: campo Risco ausente) - pula.
            stats["rows_skipped_malformed"] += 1
            continue

        if not resumo and not manejo:
            stats["rows_skipped_empty"] += 1
            continue

        if severity is None:
            severity = infer_severity("", manejo, resumo)

        drug_a = resolve_medication_name(raw_dcb, synonyms, existing_by_key)
        drug_b = resolve_medication_name(raw_farmaco, synonyms, existing_by_key)
        if drug_a.casefold() == drug_b.casefold():
            continue

        drug_a, drug_b = canonical_pair(drug_a, drug_b)
        pair_key = f"{drug_a.casefold()}||{drug_b.casefold()}"

        combined_text = " ".join([resumo, manejo])
        effects = infer_labels(combined_text, EFFECT_PATTERNS)
        systems = infer_labels(combined_text, SYSTEM_PATTERNS)
        if not systems and severity in {"major", "contraindicated"}:
            systems = ["cardiovascular"]

        candidate = {
            "pair_key": pair_key,
            "drug_a": drug_a,
            "drug_b": drug_b,
            "severity": severity,
            "action": risco,
            "mechanism": resumo,
            "recommendation": manejo,
            "effects": effects,
            "systems_affected": systems,
        }

        existing = interactions_by_pair.get(pair_key)
        if existing is None or _interaction_score(candidate) > _interaction_score(existing):
            interactions_by_pair[pair_key] = candidate

    return list(interactions_by_pair.values()), stats


def merge_formulary_medications(
    connection: sqlite3.Connection,
    formulary: dict[str, dict[str, str]],
    synonyms: dict[str, str],
    existing_by_key: dict[str, str],
    fetched_at: str,
) -> tuple[int, int]:
    created = 0
    enriched = 0

    for entry in formulary.values():
        name = resolve_medication_name(entry["display_name"], synonyms, existing_by_key)
        atc_code = entry["atc_code"]
        drug_class = atc_category(atc_code)

        row = connection.execute(
            "SELECT id, atc_code, drug_class FROM medications WHERE name = ?", (name,)
        ).fetchone()

        if row is None:
            connection.execute(
                """
                INSERT INTO medications (name, indications, drug_class, source, atc_code, updated_at)
                VALUES (?, '', ?, 'rename-uptodate', ?, ?)
                """,
                (name, drug_class, atc_code, fetched_at),
            )
            existing_by_key[strip_accents_casefold(name)] = name
            created += 1
        else:
            row_id, current_atc, current_class = row
            updates: dict[str, str] = {}
            if not current_atc and atc_code:
                updates["atc_code"] = atc_code
            if not current_class and drug_class:
                updates["drug_class"] = drug_class
            if updates:
                set_clause = ", ".join(f"{column} = ?" for column in updates)
                connection.execute(
                    f"UPDATE medications SET {set_clause} WHERE id = ?",
                    (*updates.values(), row_id),
                )
                enriched += 1

    return created, enriched


def merge_interactions(
    connection: sqlite3.Connection,
    candidates: list[dict[str, Any]],
    fetched_at: str,
) -> tuple[int, int]:
    new_count = 0
    merged_count = 0

    for candidate in candidates:
        row = connection.execute(
            """
            SELECT severity, action, mechanism, recommendation, effects_json, systems_json
            FROM interactions WHERE pair_key = ?
            """,
            (candidate["pair_key"],),
        ).fetchone()

        if row is None:
            connection.execute(
                """
                INSERT INTO interactions (
                    pair_key, drug_a_name, drug_b_name, severity, action,
                    mechanism, recommendation, effects_json, systems_json,
                    source, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'rename-uptodate', ?)
                """,
                (
                    candidate["pair_key"],
                    candidate["drug_a"],
                    candidate["drug_b"],
                    candidate["severity"],
                    candidate["action"],
                    candidate["mechanism"],
                    candidate["recommendation"],
                    json.dumps(candidate["effects"], ensure_ascii=False),
                    json.dumps(candidate["systems_affected"], ensure_ascii=False),
                    fetched_at,
                ),
            )
            new_count += 1
            continue

        (
            existing_severity,
            existing_action,
            existing_mechanism,
            existing_recommendation,
            existing_effects_json,
            existing_systems_json,
        ) = row
        existing_effects = json.loads(existing_effects_json)
        existing_systems = json.loads(existing_systems_json)

        # Reconciliação entre fontes: nunca rebaixa severidade, prefere o texto
        # mais informativo, une efeitos/sistemas. O texto de cada fonte em si
        # não é reescrito, só a escolha de qual campo prevalece no par.
        merged_severity = (
            existing_severity
            if SEVERITY_PRIORITY[existing_severity] >= SEVERITY_PRIORITY[candidate["severity"]]
            else candidate["severity"]
        )
        use_candidate_text = (
            len(candidate["mechanism"]) + len(candidate["recommendation"])
            > len(existing_mechanism) + len(existing_recommendation)
        )
        merged_mechanism = candidate["mechanism"] if use_candidate_text else existing_mechanism
        merged_recommendation = candidate["recommendation"] if use_candidate_text else existing_recommendation
        merged_action = (
            candidate["action"]
            if merged_severity == candidate["severity"] and candidate["severity"] != existing_severity
            else existing_action
        )
        merged_effects = sorted(set(existing_effects) | set(candidate["effects"]))
        merged_systems = sorted(set(existing_systems) | set(candidate["systems_affected"]))

        unchanged = (
            merged_severity == existing_severity
            and merged_mechanism == existing_mechanism
            and merged_recommendation == existing_recommendation
            and merged_effects == sorted(existing_effects)
            and merged_systems == sorted(existing_systems)
        )
        if unchanged:
            continue

        connection.execute(
            """
            UPDATE interactions SET
                severity = ?, action = ?, mechanism = ?, recommendation = ?,
                effects_json = ?, systems_json = ?, updated_at = ?
            WHERE pair_key = ?
            """,
            (
                merged_severity,
                merged_action,
                merged_mechanism,
                merged_recommendation,
                json.dumps(merged_effects, ensure_ascii=False),
                json.dumps(merged_systems, ensure_ascii=False),
                fetched_at,
                candidate["pair_key"],
            ),
        )
        merged_count += 1

    return new_count, merged_count


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Arquivo não encontrado: {xlsx_path}")

    db_path = Path(args.db)
    public_db_path = Path(args.public_db)
    synonyms = load_synonyms(Path(args.synonyms))
    fetched_at = datetime.now(UTC).replace(microsecond=0).isoformat()

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    formulary = read_formulary(workbook)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path)
    try:
        init_db(connection)

        existing_by_key = {
            strip_accents_casefold(name): name
            for (name,) in connection.execute("SELECT name FROM medications")
        }

        with connection:
            formulary_created, formulary_enriched = merge_formulary_medications(
                connection, formulary, synonyms, existing_by_key, fetched_at
            )

        # Recarrega para incluir os medicamentos do formulário recém-inseridos,
        # antes de resolver os nomes usados nas linhas de interação.
        existing_by_key = {
            strip_accents_casefold(name): name
            for (name,) in connection.execute("SELECT name FROM medications")
        }

        interaction_candidates, row_stats = read_interactions(workbook, synonyms, existing_by_key)

        with connection:
            interaction_created = 0
            for candidate in interaction_candidates:
                for name in (candidate["drug_a"], candidate["drug_b"]):
                    key = strip_accents_casefold(name)
                    if key in existing_by_key:
                        continue
                    connection.execute(
                        """
                        INSERT INTO medications (name, indications, drug_class, source, atc_code, updated_at)
                        VALUES (?, '', '', 'rename-uptodate', '', ?)
                        """,
                        (name, fetched_at),
                    )
                    existing_by_key[key] = name
                    interaction_created += 1

            interactions_new, interactions_merged = merge_interactions(
                connection, interaction_candidates, fetched_at
            )

            connection.executemany(
                """
                INSERT INTO metadata (key, value) VALUES (?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                [
                    ("rename_synced_at", fetched_at),
                    ("rename_source_file", xlsx_path.name),
                    ("rename_medications_created", str(formulary_created + interaction_created)),
                    ("rename_medications_enriched", str(formulary_enriched)),
                    ("rename_interactions_created", str(interactions_new)),
                    ("rename_interactions_merged", str(interactions_merged)),
                ],
            )

        publish_frontend_database(connection, public_db_path)
    finally:
        connection.close()

    print("Import do RENAME concluído.")
    print(f"  Linhas de interação lidas: {row_stats['rows_seen']}")
    print(f"  Linhas puladas (colunas deslocadas na origem): {row_stats['rows_skipped_malformed']}")
    print(f"  Linhas puladas (sem texto clínico): {row_stats['rows_skipped_empty']}")
    print(f"  Medicamentos novos: {formulary_created + interaction_created}")
    print(f"  Medicamentos existentes enriquecidos (ATC/classe): {formulary_enriched}")
    print(f"  Interações novas: {interactions_new}")
    print(f"  Interações já existentes, mescladas com o RENAME: {interactions_merged}")
    print(f"SQLite local: {db_path}")
    print(f"SQLite público: {public_db_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
